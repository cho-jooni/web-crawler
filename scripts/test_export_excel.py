"""scripts/test_export_excel.py"""
import os
from openpyxl import load_workbook
from export_excel import export_to_excel


def test_export_creates_file(tmp_path):
    data = [
        {"상품명": "테스트 상품 A", "가격": 15000, "리뷰수": 42},
        {"상품명": "테스트 상품 B", "가격": 23000, "리뷰수": 17},
    ]
    filepath = str(tmp_path / "test_output.xlsx")
    result = export_to_excel(data, filepath)
    assert os.path.exists(result)


def test_export_correct_content(tmp_path):
    data = [
        {"상품명": "상품A", "가격": 10000},
        {"상품명": "상품B", "가격": 20000},
    ]
    filepath = str(tmp_path / "test.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    # Header
    assert ws.cell(1, 1).value == "상품명"
    assert ws.cell(1, 2).value == "가격"
    # Data
    assert ws.cell(2, 1).value == "상품A"
    assert ws.cell(2, 2).value == 10000
    assert ws.cell(3, 1).value == "상품B"
    assert ws.max_row == 3  # header + 2 rows


def test_export_header_bold(tmp_path):
    data = [{"col1": "val"}]
    filepath = str(tmp_path / "test.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.cell(1, 1).font.bold is True


def test_export_auto_filter(tmp_path):
    data = [{"a": 1, "b": 2}]
    filepath = str(tmp_path / "test.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.auto_filter.ref is not None


def test_export_empty_data(tmp_path):
    filepath = str(tmp_path / "empty.xlsx")
    result = export_to_excel([], filepath)
    assert result is None  # No file created for empty data


def test_export_korean_encoding(tmp_path):
    data = [{"이름": "홍길동", "주소": "서울시 강남구"}]
    filepath = str(tmp_path / "korean.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.cell(2, 1).value == "홍길동"


# ── 산출물 무결성 ──
# 엑셀은 파이프라인의 마지막 단계다. 여기서 조용히 열이 빠지거나 크래시가 나면
# 오래 걸린 수집 결과가 통째로 날아간다.

def test_headers_are_union_of_all_rows(tmp_path):
    """뒤쪽 행에만 있는 필드가 사라지면 안 된다 — 스크래핑에서는 정상적인 상황이다."""
    data = [
        {"상품명": "A", "가격": 100},
        {"상품명": "B", "가격": 200, "할인가": 150},
        {"상품명": "C", "가격": 300, "품절": True},
    ]
    filepath = str(tmp_path / "union.xlsx")
    export_to_excel(data, filepath)
    ws = load_workbook(filepath).active
    assert [c.value for c in ws[1]] == ["상품명", "가격", "할인가", "품절"]
    assert ws.cell(3, 3).value == 150
    assert ws.cell(4, 4).value is True
    assert ws.cell(2, 3).value is None      # 없는 값은 빈 칸으로 남는다


def test_list_value_is_flattened(tmp_path):
    filepath = str(tmp_path / "list.xlsx")
    export_to_excel([{"이미지": ["a.jpg", "b.jpg"]}], filepath)
    assert load_workbook(filepath).active.cell(2, 1).value == "a.jpg, b.jpg"


def test_dict_value_is_serialized_as_json(tmp_path):
    filepath = str(tmp_path / "dict.xlsx")
    export_to_excel([{"옵션": {"색상": "빨강"}}], filepath)
    assert load_workbook(filepath).active.cell(2, 1).value == '{"색상": "빨강"}'


def test_nested_list_value_is_serialized_as_json(tmp_path):
    filepath = str(tmp_path / "nested.xlsx")
    export_to_excel([{"옵션": [{"색상": "빨강"}]}], filepath)
    assert load_workbook(filepath).active.cell(2, 1).value == '[{"색상": "빨강"}]'


def test_control_characters_are_stripped(tmp_path):
    """스크랩한 본문에 제어문자가 섞이는 일은 흔하다. 남겨두면 저장 시 죽는다."""
    filepath = str(tmp_path / "ctrl.xlsx")
    export_to_excel([{"본문": "리뷰\x0b본문\x00끝"}], filepath)
    assert load_workbook(filepath).active.cell(2, 1).value == "리뷰본문끝"


def test_overlong_text_is_truncated_to_cell_limit(tmp_path):
    """셀 상한을 넘기면 크래시는 안 나지만 엑셀이 파일을 못 연다 — 잘라서 표시를 남긴다."""
    filepath = str(tmp_path / "long.xlsx")
    export_to_excel([{"본문": "가" * 40000}], filepath)
    value = load_workbook(filepath).active.cell(2, 1).value
    assert len(value) == 32767
    assert value.endswith("…(잘림)")


def test_non_dict_rows_are_skipped_not_fatal(tmp_path):
    filepath = str(tmp_path / "mixed.xlsx")
    export_to_excel([{"a": 1}, "쓰레기", {"a": 2}], filepath)
    ws = load_workbook(filepath).active
    assert ws.cell(2, 1).value == 1
    assert ws.cell(4, 1).value == 2


def test_rows_without_any_keys_produce_no_file(tmp_path):
    filepath = str(tmp_path / "nokeys.xlsx")
    assert export_to_excel([{}, {}], filepath) is None
