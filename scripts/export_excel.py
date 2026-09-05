"""scripts/export_excel.py — 수집 데이터를 엑셀 파일로 출력"""
import json
import os
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 엑셀 셀 하나에 들어갈 수 있는 문자 수 상한(스펙). 넘기면 openpyxl 은 그대로 쓰지만
# 엑셀이 파일을 "읽을 수 없는 내용" 으로 판정한다 — 즉 크래시 없이 산출물만 깨진다.
_CELL_TEXT_LIMIT = 32767
_TRUNCATED_MARK = "…(잘림)"


def _collect_headers(data: list[dict]) -> list[str]:
    """모든 행의 키를 합집합으로 모은다 — 처음 나온 순서를 유지한다.

    첫 행의 키만 헤더로 쓰면, 뒤쪽 행에만 있는 필드가 **경고 없이 통째로 사라진다.**
    스크래핑 데이터에서 이건 예외가 아니라 정상이다 — 할인가·옵션·품절표시처럼
    일부 항목에만 붙는 필드가 흔하다. 조용한 유실은 크래시보다 나쁘다.
    """
    headers: dict[str, None] = {}
    for row in data:
        if isinstance(row, dict):
            for key in row:
                headers.setdefault(key, None)
    return list(headers)


def _cell_value(value):
    """엑셀 셀에 넣을 수 있는 형태로 바꾼다.

    수집 데이터에는 엑셀이 그대로 받지 못하는 값이 섞인다. 마지막 단계에서
    ValueError 로 죽으면 오래 걸린 수집 결과가 통째로 날아가므로 여기서 흡수한다.

      · list/dict  — 이미지 목록·옵션 같은 중첩 값. 스칼라 리스트는 읽기 좋게 ", " 로,
                     그 외에는 JSON 으로 편다. (openpyxl 은 ValueError 를 낸다)
      · 제어문자    — 스크랩한 본문에 자주 섞인다. 남겨두면 IllegalCharacterError 로 죽는다
      · 초장문      — 셀 상한을 넘으면 파일이 열리지 않으므로 자르고 표시를 남긴다
    """
    if value is None or isinstance(value, (int, float, bool)):
        return value

    if isinstance(value, (list, tuple, set)):
        items = list(value)
        if all(isinstance(v, (str, int, float, bool)) or v is None for v in items):
            value = ", ".join("" if v is None else str(v) for v in items)
        else:
            value = json.dumps(items, ensure_ascii=False, default=str)
    elif isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False, default=str)
    elif not isinstance(value, str):
        # datetime 등 openpyxl 이 아는 타입은 그대로 두고, 모르는 것만 문자열로 만든다.
        if type(value).__module__ == "datetime":
            return value
        value = str(value)

    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(value) > _CELL_TEXT_LIMIT:
        value = value[:_CELL_TEXT_LIMIT - len(_TRUNCATED_MARK)] + _TRUNCATED_MARK
    return value


def export_to_excel(data: list[dict], filepath: str, sheet_name: str = "수집 데이터") -> str | None:
    """데이터 리스트를 엑셀 파일로 저장.

    Args:
        data: 딕셔너리 리스트. **모든 행**의 키가 헤더가 됨(합집합, 첫 등장 순서).
        filepath: 저장할 .xlsx 파일 경로.
        sheet_name: 시트 이름 (기본값: "수집 데이터").

    Returns:
        저장된 파일 경로. 데이터가 없으면 None.
    """
    if not data:
        return None

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = _collect_headers(data)
    if not headers:
        return None

    # Header row (bold)
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=_cell_value(header))
        cell.font = bold_font

    # Data rows
    for row_idx, item in enumerate(data, 2):
        if not isinstance(item, dict):
            continue
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(item.get(header)))

    # Auto column width
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in range(2, len(data) + 2):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    wb.save(filepath)
    return filepath
